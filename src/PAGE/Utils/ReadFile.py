import openpyxl

class Excel_File:

    def get_Row(file, sheet_name, row_num):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheet_name]
        row = sheet[row_num]
        return row

    def get_Col(file, sheet_name, col_num):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheet_name]
        col = sheet[col_num]
        return col

    def read_data(file, sheet_name, row_num, col_num):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheet_name]
        return sheet.cell(row=row_num, column=col_num).value

    def write_data(file, sheet_name, row_num, col_num, data):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.get_sheet_by_name(sheet_name)
        sheet.cell(row=row_num, column=col_num).value = data
        workbook.save(file)